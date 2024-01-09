
from PyQt5.QtWidgets import QFileDialog
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
import sys
from PyQt5 import QtCore, QtGui, QtWidgets
import pandas as pd

class Ui_Form(object):

    # defining the global variables
    windspeed_column = None
    start_date = None
    height_of_windspeed = None
    alpha = None
    turbine_name = None
    wind_cell = None
    power_cell = None
    hub_height = None
    dataframe = None
    cutin = None
    cutout = None
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(1014, 536)
        self.gridLayout = QtWidgets.QGridLayout(Form)
        self.gridLayout.setObjectName("gridLayout")
        self.verticalLayout_4 = QtWidgets.QVBoxLayout()
        self.verticalLayout_4.setObjectName("verticalLayout_4")
        self.pushButton_data = QtWidgets.QPushButton(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.pushButton_data.setFont(font)
        self.pushButton_data.setObjectName("pushButton_data")

        # connection for uploading weather data
        self.pushButton_data.clicked.connect(self.upload_dataframe)

        self.verticalLayout_4.addWidget(self.pushButton_data)
        self.lineEdit_column = QtWidgets.QLineEdit(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lineEdit_column.setFont(font)
        self.lineEdit_column.setObjectName("lineEdit_column")
        self.verticalLayout_4.addWidget(self.lineEdit_column)
        self.lineEdit_date = QtWidgets.QLineEdit(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lineEdit_date.setFont(font)
        self.lineEdit_date.setObjectName("lineEdit_date")
        self.verticalLayout_4.addWidget(self.lineEdit_date)
        self.lineEdit_height_wind = QtWidgets.QLineEdit(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lineEdit_height_wind.setFont(font)
        self.lineEdit_height_wind.setObjectName("lineEdit_height_wind")
        self.verticalLayout_4.addWidget(self.lineEdit_height_wind)
        self.lineEdit_alpha = QtWidgets.QLineEdit(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lineEdit_alpha.setFont(font)
        self.lineEdit_alpha.setObjectName("lineEdit_alpha")
        self.verticalLayout_4.addWidget(self.lineEdit_alpha)
        self.gridLayout.addLayout(self.verticalLayout_4, 0, 1, 1, 1)
        self.pushButton_simulate = QtWidgets.QPushButton(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_simulate.setFont(font)
        self.pushButton_simulate.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        self.pushButton_simulate.setObjectName("pushButton_simulate")

        # connection to the button
        self.pushButton_simulate.clicked.connect(self.solve_wind_energy)

        self.gridLayout.addWidget(self.pushButton_simulate, 2, 3, 1, 1)
        self.verticalLayout_3 = QtWidgets.QVBoxLayout()
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.label_turbine = QtWidgets.QLabel(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_turbine.setFont(font)
        self.label_turbine.setObjectName("label_turbine")
        self.verticalLayout_3.addWidget(self.label_turbine)
        self.label_wind_cell = QtWidgets.QLabel(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_wind_cell.setFont(font)
        self.label_wind_cell.setObjectName("label_wind_cell")
        self.verticalLayout_3.addWidget(self.label_wind_cell)
        self.label_power_cell = QtWidgets.QLabel(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_power_cell.setFont(font)
        self.label_power_cell.setObjectName("label_power_cell")
        self.verticalLayout_3.addWidget(self.label_power_cell)
        self.label_Hub = QtWidgets.QLabel(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_Hub.setFont(font)
        self.label_Hub.setObjectName("label_Hub")
        self.verticalLayout_3.addWidget(self.label_Hub)
        self.label_cutin = QtWidgets.QLabel(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_cutin.setFont(font)
        self.label_cutin.setObjectName("label_cutin")
        self.verticalLayout_3.addWidget(self.label_cutin)
        self.label_cutout = QtWidgets.QLabel(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_cutout.setFont(font)
        self.label_cutout.setObjectName("label_cutout")
        self.verticalLayout_3.addWidget(self.label_cutout)
        self.gridLayout.addLayout(self.verticalLayout_3, 0, 3, 1, 1)
        self.pushButton_input = QtWidgets.QPushButton(Form)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Minimum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_input.sizePolicy().hasHeightForWidth())
        self.pushButton_input.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(False)
        font.setWeight(50)
        self.pushButton_input.setFont(font)
        self.pushButton_input.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        self.pushButton_input.setObjectName("pushButton_input")

        # connection to the button
        self.pushButton_input.clicked.connect(self.wind_column)
        self.pushButton_input.clicked.connect(self.start_date)
        self.pushButton_input.clicked.connect(self.Height_measured_windspeed)
        self.pushButton_input.clicked.connect(self.alpha)
        self.pushButton_input.clicked.connect(self.Turbine)
        self.pushButton_input.clicked.connect(self.wind_cell)
        self.pushButton_input.clicked.connect(self.power_cell)
        self.pushButton_input.clicked.connect(self.hub_height)
        self.pushButton_input.clicked.connect(self.cutin)
        self.pushButton_input.clicked.connect(self.cutout)

        self.gridLayout.addWidget(self.pushButton_input, 2, 1, 1, 1)
        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.label_weather_data = QtWidgets.QLabel(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_weather_data.setFont(font)
        self.label_weather_data.setObjectName("label_weather_data")
        self.verticalLayout_2.addWidget(self.label_weather_data)
        self.label_column = QtWidgets.QLabel(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_column.setFont(font)
        self.label_column.setObjectName("label_column")
        self.verticalLayout_2.addWidget(self.label_column)
        self.label_date = QtWidgets.QLabel(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_date.setFont(font)
        self.label_date.setObjectName("label_date")
        self.verticalLayout_2.addWidget(self.label_date)
        self.label_height_wind = QtWidgets.QLabel(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_height_wind.setFont(font)
        self.label_height_wind.setObjectName("label_height_wind")
        self.verticalLayout_2.addWidget(self.label_height_wind)
        self.label_alpha = QtWidgets.QLabel(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_alpha.setFont(font)
        self.label_alpha.setObjectName("label_alpha")
        self.verticalLayout_2.addWidget(self.label_alpha)
        self.gridLayout.addLayout(self.verticalLayout_2, 0, 0, 1, 1)
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.lineEdit_turbine = QtWidgets.QLineEdit(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lineEdit_turbine.setFont(font)
        self.lineEdit_turbine.setObjectName("lineEdit_turbine")
        self.verticalLayout.addWidget(self.lineEdit_turbine)
        self.lineEdit_wind_cell = QtWidgets.QLineEdit(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lineEdit_wind_cell.setFont(font)
        self.lineEdit_wind_cell.setObjectName("lineEdit_wind_cell")
        self.verticalLayout.addWidget(self.lineEdit_wind_cell)
        self.lineEdit_power_cell = QtWidgets.QLineEdit(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lineEdit_power_cell.setFont(font)
        self.lineEdit_power_cell.setObjectName("lineEdit_power_cell")
        self.verticalLayout.addWidget(self.lineEdit_power_cell)
        self.lineEdit_hub = QtWidgets.QLineEdit(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lineEdit_hub.setFont(font)
        self.lineEdit_hub.setObjectName("lineEdit_hub")
        self.verticalLayout.addWidget(self.lineEdit_hub)
        self.lineEdit_cutin = QtWidgets.QLineEdit(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lineEdit_cutin.setFont(font)
        self.lineEdit_cutin.setObjectName("lineEdit_cutin")
        self.verticalLayout.addWidget(self.lineEdit_cutin)
        self.lineEdit_cutout = QtWidgets.QLineEdit(Form)
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lineEdit_cutout.setFont(font)
        self.lineEdit_cutout.setObjectName("lineEdit_cutout")
        self.verticalLayout.addWidget(self.lineEdit_cutout)
        self.gridLayout.addLayout(self.verticalLayout, 0, 4, 1, 1)
        self.plainTextEdit = QtWidgets.QPlainTextEdit(Form)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Maximum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.plainTextEdit.sizePolicy().hasHeightForWidth())
        self.plainTextEdit.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.plainTextEdit.setFont(font)
        self.plainTextEdit.setObjectName("plainTextEdit")
        self.gridLayout.addWidget(self.plainTextEdit, 3, 0, 1, 5)
        spacerItem = QtWidgets.QSpacerItem(13, 13, QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout.addItem(spacerItem, 0, 2, 1, 1)

        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Form"))
        self.pushButton_data.setText(_translate("Form", "Upload"))
        self.lineEdit_column.setPlaceholderText(_translate("Form", "Ex: wind_speed"))
        self.lineEdit_date.setPlaceholderText(_translate("Form", "Ex: 2005-01-01"))
        self.lineEdit_height_wind.setPlaceholderText(_translate("Form", "in meters"))
        self.lineEdit_alpha.setPlaceholderText(_translate("Form", "Ex: 0.143"))
        self.pushButton_simulate.setText(_translate("Form", "Simulate"))
        self.label_turbine.setText(_translate("Form", "Turbine name:"))
        self.label_wind_cell.setText(_translate("Form", "wind speed cell number:"))
        self.label_power_cell.setText(_translate("Form", "Power values cell number:"))
        self.label_Hub.setText(_translate("Form", "Hub height:"))
        self.label_cutin.setText(_translate("Form", "Cut in speed:"))
        self.label_cutout.setText(_translate("Form", "Cut out speed:"))
        self.pushButton_input.setText(_translate("Form", "          Load inputs           "))
        self.label_weather_data.setText(_translate("Form", "Weather data:"))
        self.label_column.setText(_translate("Form", "wind speed column name:"))
        self.label_date.setText(_translate("Form", "start date of data:"))
        self.label_height_wind.setText(_translate("Form", "Height of measured wind speed:"))
        self.label_alpha.setText(_translate("Form", "Terrain roughness coefficient (α):"))
        self.lineEdit_turbine.setPlaceholderText(_translate("Form", "from datasheet"))
        self.lineEdit_wind_cell.setPlaceholderText(_translate("Form", "from datasheet (Ex: B2)"))
        self.lineEdit_power_cell.setPlaceholderText(_translate("Form", "from datasheet (Ex: C2)"))
        self.lineEdit_hub.setPlaceholderText(_translate("Form", "in meters"))
        self.lineEdit_cutin.setPlaceholderText(_translate("Form", "in m/sec"))
        self.lineEdit_cutout.setPlaceholderText(_translate("Form", "in m/sec"))


    # function for weather data
    def upload_dataframe(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_name, _ = QFileDialog.getOpenFileName(None, "Open CSV File", "",
                                                   "CSV Files (*.csv);;All Files (*)",
                                                   options=options)
        if file_name:
            # Read the CSV file into a DataFrame
            try:
                dataframe = pd.read_csv(file_name)
                self.dataframe = dataframe
                #print("weather data uploaded")
                self.plainTextEdit.appendPlainText("weather data uploaded")
                #print(dataframe)
                return dataframe
            except Exception as e:
                #print("Error in uploading selected file")
                self.plainTextEdit.appendPlainText("Error uploading the selected file")

    def wind_column(self):
        windspeed_column = self.lineEdit_column.text()
        self.windspeed_column = windspeed_column
        #print(windspeed_column)
        self.plainTextEdit.appendPlainText(f"Wind speed column  :{windspeed_column}")
        return windspeed_column

    def start_date(self):
        start_date = self.lineEdit_date.text()
        self.start_date = start_date
        #print(start_date)
        self.plainTextEdit.appendPlainText(f"Start date :{start_date}")
        return start_date

    def Height_measured_windspeed(self):
        height_of_windspeed = self.lineEdit_height_wind.text()

        try:
            height_of_windspeed = float(height_of_windspeed)
            self.height_of_windspeed = height_of_windspeed
            #print(height_of_windspeed)
            self.plainTextEdit.appendPlainText(f"Wind speed measured height :{height_of_windspeed}")
            return height_of_windspeed
        except ValueError:
            #print("invalid value, please enter a valid number for height")
            self.plainTextEdit.appendPlainText("invalid value, please enter a valid number for height")

    def alpha(self):
        alpha = self.lineEdit_alpha.text()

        try:
            alpha = float(alpha)
            self.alpha = alpha
            #print(alpha)
            self.plainTextEdit.appendPlainText(f"Alpha :{alpha}")
            return alpha
        except ValueError:
            #print("invalid value, please enter a valid number for alpha")
            self.plainTextEdit.appendPlainText("invalid value, please enter a valid number for alpha")

    def Turbine(self):
        turbine_name = self.lineEdit_turbine.text()
        self.turbine_name = turbine_name
        #print(turbine_name)
        self.plainTextEdit.appendPlainText(f"Turbine name :{turbine_name}")
        return turbine_name

    def wind_cell(self):
        wind_cell = self.lineEdit_wind_cell.text()
        self.wind_cell = wind_cell
        #print(wind_cell)
        self.plainTextEdit.appendPlainText(f"Wind speed Cell number :{wind_cell}")
        return wind_cell

    def power_cell(self):
        power_cell = self.lineEdit_power_cell.text()
        self.power_cell = power_cell
        #print(power_cell)
        self.plainTextEdit.appendPlainText(f"Power curve Cell number :{power_cell}")
        return power_cell

    def hub_height(self):
        hub_height = self.lineEdit_hub.text()
        try:
            hub_height = float(hub_height)
            self.hub_height = hub_height
            #print(hub_height)
            self.plainTextEdit.appendPlainText(f"Hub height :{hub_height}")
            return hub_height
        except ValueError:
            #print("invalid value, please enter a valid number for hub height")
            self.plainTextEdit.appendPlainText("invalid value, please enter a valid number for hub height")

    def cutin(self):
        cutin = self.lineEdit_cutin.text()

        try:
            cutin = float(cutin)
            self.cutin = cutin
            #print(cutin)
            self.plainTextEdit.appendPlainText(f"Cut-in speed :{cutin}")
            return cutin
        except ValueError:
            #print("invalid value, please enter a number for cut-in speed")
            self.plainTextEdit.appendPlainText("invalid value, please enter a number for cut-in speed")

    def cutout(self):
        cutout = self.lineEdit_cutout.text()

        try:
            cutout = float(cutout)
            self.cutout = cutout
            #print(cutout)
            self.plainTextEdit.appendPlainText(f"Cut-out speed :{cutout}")
            return cutout
        except ValueError:
            #print("invalid value, please enter a number for cut-out speed")
            self.plainTextEdit.appendPlainText("invalid value, please enter a number for cut-out speed")

    def solve_wind_energy(self):

        windspeed_column_name = self.windspeed_column
        start_date_of_data = self.start_date
        height_of_measured_wind_speed = self.height_of_windspeed
        alpha_value = self.alpha
        turbine_selected = self.turbine_name
        windspeed_cell = self.wind_cell
        power_curve_cell = self.power_cell
        hub_height_input = self.hub_height
        cut_in = self.cutin
        cut_out = self.cutout
        wind_speed_data = self.dataframe

        # renaming the column name of wind speed to wind_speed
        wind_speed_data.rename(columns={windspeed_column_name: 'wind_speed'}, inplace=True)
        wind_speed_data['DatetimeIndex'] = pd.date_range(start=start_date_of_data, periods=len(wind_speed_data),
                                                         freq='H')

        # Loading the power curve data from an Excel sheet from the cell
        wb = load_workbook('power_curve.xlsx', data_only=True)
        sheet = wb.active
        power_curve_data = list(sheet.iter_rows(values_only=True))

        # Extracting the 'wind_speed' values as a string
        wind_speed_str = sheet[windspeed_cell].value
        # Converting the string to a list of float values
        wind_speeds = [float(val) for val in wind_speed_str.strip('[]').split(',')]

        # Extracting the 'wind_speed' values as a string
        power_output_str = sheet[power_curve_cell].value
        # Converting the string to a list of float values
        power_outputs = [float(val) for val in power_output_str.strip('[]').split(',')]

        # Set the rated power of the turbine as the max value from power curve
        rated_power = max(power_outputs)

        # Cut-in and cut-out wind speeds
        cut_in_speed = cut_in
        cut_out_speed = cut_out

        Hub_height = hub_height_input  # height of hub in meters
        alpha = alpha_value  # terrain roughness coefficient
        measured_wind_speed_height = height_of_measured_wind_speed  # measured height of wind speed in meters

        # Initializing the df to store the calculated energy yield
        hourly_energy_yield = pd.DataFrame(index=wind_speed_data.index, columns=['EnergyYield_kWh'])

        # Iterating through the hourly wind speed data
        for index, row in wind_speed_data.iterrows():
            wind_speed_at_measured = row['wind_speed']

            # calculating wind speed at hub
            wind_speed_at_hub = wind_speed_at_measured * (Hub_height / measured_wind_speed_height) ** alpha

            # if wind speed is within the operational range
            if cut_in_speed <= wind_speed_at_hub <= cut_out_speed:
                # linear interpolation to estimate power output
                power_output = np.interp(wind_speed_at_hub, wind_speeds, power_outputs)
            else:
                power_output = 0.0  # Turbine is at halt

            # Calculate energy produced in this hour (kWh)
            energy_produced = power_output
            hourly_energy_yield.loc[index, 'EnergyYield'] = energy_produced

        df = hourly_energy_yield
        df['DatetimeIndex'] = pd.date_range(start=start_date_of_data, periods=len(wind_speed_data), freq='H')

        # Export the hourly energy yield with datetime index to a CSV file
        #df.to_csv('hourly_energy_yield.csv')

        annual_energy_yield = df['EnergyYield'].sum()
        scaled_data = hourly_energy_yield['EnergyYield'] / hourly_energy_yield['EnergyYield'].max()
        #print(annual_energy_yield)
        self.plainTextEdit.setPlainText(f"Annual energy output :{annual_energy_yield}")
        scaled_data.index =pd.date_range(start=start_date_of_data, periods=len(wind_speed_data), freq='H')
        scaled_data.to_csv('normalised_wind_output.csv')
        #print("Normalised output is saved into current directory")
        self.plainTextEdit.appendPlainText(f"Annual energy peak :{hourly_energy_yield['EnergyYield'].max()}")
        self.plainTextEdit.appendPlainText("Normalised output is saved to current directory")

        # Create a plot of energy yield
        plt.figure
        plt.plot(scaled_data, label=turbine_selected)
        plt.xlabel('Datetime')
        plt.ylabel('Normalised power')
        plt.title('Normalised Hourly Wind Turbine Energy Yield')
        plt.grid(True)
        plt.show()


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Form = QtWidgets.QWidget()
    ui = Ui_Form()
    ui.setupUi(Form)
    Form.show()
    sys.exit(app.exec_())
